"""
CSV to Excel Automation Script
-------------------------------------------------

Loads data from CSV file into Excel sheet.

Features:
- Dynamic configuration using config.py
- Includes first CSV row as data
- Supports dynamic sheet names
- Supports dynamic start row
- Supports dynamic target columns
- Supports prefixes for specific columns

Requirements:
    pip install pandas openpyxl
"""

import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from config import (
    DATA_FOLDER,
    CSV_FILE_NAME,
    EXCEL_FILE_NAME,
    SHEET_NAME,
    START_ROW,
    TARGET_COLUMNS,
    COLUMN_PREFIXES
)


# =========================================================
# BUILD FILE PATHS
# =========================================================

csv_path = DATA_FOLDER / CSV_FILE_NAME

excel_path = DATA_FOLDER / EXCEL_FILE_NAME


# =========================================================
# DEBUG INFORMATION
# =========================================================

print("======================================")
print("CSV Path:", csv_path)
print("Excel Path:", excel_path)
print("CSV Exists:", csv_path.exists())
print("Excel Exists:", excel_path.exists())
print("======================================")


# =========================================================
# VALIDATE FILES
# =========================================================

if not csv_path.exists():
    raise FileNotFoundError(
        f"CSV file not found:\n{csv_path}"
    )

if not excel_path.exists():
    raise FileNotFoundError(
        f"Excel file not found:\n{excel_path}"
    )


# =========================================================
# READ CSV FILE
# =========================================================

print(f"Reading CSV file:\n{csv_path}")

# header=None ensures first row is included as data
df = pd.read_csv(csv_path, header=None)

if df.empty:
    raise Exception("CSV file is empty.")

required_column_count = len(TARGET_COLUMNS)

if len(df.columns) < required_column_count:
    raise Exception(
        f"CSV must contain at least "
        f"{required_column_count} columns."
    )

# Only use required columns
df = df.iloc[:, :required_column_count]


# =========================================================
# LOAD EXCEL WORKBOOK
# =========================================================

print(f"Opening Excel workbook:\n{excel_path}")

workbook = load_workbook(excel_path)

if SHEET_NAME not in workbook.sheetnames:
    raise Exception(
        f"Sheet '{SHEET_NAME}' not found in workbook."
    )

worksheet = workbook[SHEET_NAME]


# =========================================================
# WRITE DATA INTO EXCEL
# =========================================================

print("Writing data into Excel...")

for row_offset, row_data in enumerate(df.values):

    excel_row = START_ROW + row_offset

    for col_offset, value in enumerate(row_data):

        excel_column_letter = TARGET_COLUMNS[col_offset]

        excel_column_number = column_index_from_string(
            excel_column_letter
        )

        # =================================================
        # APPLY PREFIX IF CONFIGURED
        # =================================================

        prefix = COLUMN_PREFIXES.get(
            excel_column_letter
        )

        if prefix:
            value = f"{prefix}{value}"

        worksheet.cell(
            row=excel_row,
            column=excel_column_number,
            value=value
        )


# =========================================================
# SAVE OUTPUT FILE
# =========================================================

output_file = (
    DATA_FOLDER /
    "CRMTeam StatusReport_JUNE2026_UPDATED.xlsx"
)

workbook.save(output_file)


# =========================================================
# SUCCESS MESSAGE
# =========================================================

print("======================================")
print("Data inserted successfully.")
print("Output File:")
print(output_file)
print("======================================")